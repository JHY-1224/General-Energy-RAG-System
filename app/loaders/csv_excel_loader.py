from __future__ import annotations

import csv
from pathlib import Path

from .base_loader import BaseLoader, LoadedDocument, decode_text


def _rows_to_markdown(rows: list[list[object]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [[str(cell) for cell in row] + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in normalized[1:])
    return "\n".join(lines)


class CsvExcelLoader(BaseLoader):
    def load(self, path: str | Path) -> list[LoadedDocument]:
        source = Path(path)
        if source.suffix.lower() == ".csv":
            rows = list(csv.reader(decode_text(source).splitlines()))
            parser = "csv"
        else:
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(source, read_only=True, data_only=True)
                rows = []
                for sheet in workbook.worksheets:
                    rows.append([f"Sheet: {sheet.title}"])
                    rows.extend([list(row) for row in sheet.iter_rows(values_only=True)])
                parser = "openpyxl"
            except ModuleNotFoundError:
                rows = [["Excel 文件", source.name], ["提示", "安装 openpyxl 后可提取字段和表格"]]
                parser = "placeholder"
        return [LoadedDocument(content=_rows_to_markdown(rows), source=source.name, metadata={"format": source.suffix.lower().lstrip("."), "parser": parser, "category": "Table"})]
